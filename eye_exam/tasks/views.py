from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from .models import Test, Question, Student, Test_result, Answer
from django.http import JsonResponse
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import json
from collections import defaultdict
from django.http import FileResponse

SPACER = ';'
excel_stats = bytearray()

def test_list(request, test_id):
    test = Test.objects.filter(pk=test_id)
    questions = Question.objects.filter(test=test_id)
    answers = Answer.objects.all()

    result_questions = []
    all_correct_ans = []
    get_test_obj = Test.objects.get(pk=test_id)

    students = Student.objects.filter(grade_num=get_test_obj.test_grade_num, grade_letter=get_test_obj.test_grade_letter)

    for q in questions:
        cnt = 0
        array = []
        numbers = []
        ans = []
        for j in answers:
            if j.question.pk == q.pk:
                ans.append(j)

        for i in ans:
            cnt += 1
            i = str(cnt) + ") " + str(i)
            array.append(i)
            numbers.append(cnt)

        for i in ans:
            if i.is_correct == True:
                all_correct_ans.append(str(ans.index(i)))
        result_questions.append({
            "name": q.text,
            "number": numbers,
            "answers":array,
        })

    get_test_obj.test_correct_answers = SPACER.join(all_correct_ans)
    get_test_obj.save()

    return render(request, 'tasks/test_list.html', {'test': test, 'questions': result_questions, 'students': students})


@csrf_exempt
def checkAnswer(request):
    if request.method == 'POST':
        body_unicode = request.body.decode('utf-8')
        body = json.loads(body_unicode)
        students_answers = body['student_answers']
        student_id = body['student_id']
        test_id = body['test_id']
        test = Test.objects.get(pk=test_id)
        correct_answers = test.test_correct_answers
        new_correct_answers = correct_answers.replace(';','')
        answers = []
        st_ans = []
        stats = []
        score = 0
        for i in new_correct_answers:
            answers.append(int(i))
        for key, value in students_answers.items():
            st_ans.append(value)
        for i in range(len(answers)):
            if answers[i] == st_ans[i]:
                score += 1
                stats.append('1')
            else:
                score += 0
                stats.append('0')
        get_test = Test.objects.get(pk=test_id)
        get_student = Student.objects.get(pk=student_id)
        get_test_res = Test_result.objects.filter(test=get_test, student=get_student.name + ' ' + get_student.surname)
        if len(get_test_res) != 0:
            got_test_res = Test_result.objects.get(test=get_test, student=get_student.name + ' ' + get_student.surname)
            got_test_res.score = score
            got_test_res.max_score = len(answers)
            got_test_res.stats = SPACER.join(stats)
            got_test_res.save(update_fields=['score', 'max_score', 'stats'])
        else:
            reg = Test_result(test=get_test,
                                student_id=student_id,
                                teacher=get_test.created_by,
                                student=get_student.name + ' ' + get_student.surname,
                                grade=str(get_student.grade_num) + ' ' + get_student.grade_letter,
                                score=score,
                                max_score=len(answers),
                                stats=SPACER.join(stats),
                                slug=get_test.slug)
            reg.save()
        return JsonResponse({'score': score, 'max_score': len(answers), 'student': get_student.surname + ' ' + get_student.name})



def generate_excel(request, slug):
    get_res = Test_result.objects.filter(slug=slug)
    array = defaultdict(list)
    for i in get_res:
        res_stats = []
        stats = i.stats.split(';')
        get_test = Test.objects.get(pk=i.test_id)
        array['Teacher'] = str(get_test.created_by)
        array['Test'] = get_test.text
        for j in stats:
            j = int(j)
            res_stats.append(j)
        array['Students'].append({"Result": res_stats, 'Student': i.student})

    data = array
    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика"
    x = 2
    ws.cell(row=len(data['Students'][0]['Result']) + 2, column=1, value='РЕЗУЛЬТАТ')
    [ws.cell(row=res + 2, column=1, value="№ " + str(res + 1)) for res in range(len(data['Students'][0]['Result']))]
    for student in data['Students']:
        ws.cell(row=1, column=x, value=student['Student'])
        y = 2
        for res in student['Result']:
            ws.cell(row=y, column=x, value=res)
            y += 1
        letter = ws.cell(row=2, column=x).column_letter
        c = ws.cell(row=y, column=x, value=f'=SUM({letter}2:{letter}{y - 1})')
        c.fill = PatternFill('solid', fgColor="DDDDDD")
        x += 1
    letter1 = ws.cell(row=2, column=x + 1).column_letter
    letter2 = ws.cell(row=2, column=x + 2).column_letter
    ws.merge_cells(f'{letter1}2:{letter2}2')
    ws.cell(row=2, column=x + 1, value=data['Teacher'])

    ws.merge_cells(f'{letter1}4:{letter2}4')
    ws.cell(row=4, column=x + 1, value=data['Test'])

    ws.merge_cells(f'{letter1}6:{letter2}6')
    ws.cell(row=6, column=x + 1, value='Средний балл по классу')

    n = len(data['Students'][0]['Result']) + 2
    ws.cell(row=6, column=x + 3, value=f'=AVERAGE(A{n}:{letter}{n})')

    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    return FileResponse(virtual_workbook, as_attachment=True, filename='Статистика.xlsx')
